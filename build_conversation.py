# -*- coding: utf-8 -*-
"""
MedSnap 对话构建与数据处理模块
提供 OCR 文本结构化、结构化数据统计分析、Excel 导出等功能。
"""

import os
import json
import numpy as np
import pandas as pd
from datetime import datetime

from model import client, MODEL_NAME, parse_ai_response
from desensitizer import desensitize_text
from template import CATEGORY_CONFIGS

# ========== OCR 文本 → LLM 结构化 ==========

def extract_from_ocr_text(ocr_text, ai_prompt):
    """
    将本地 OCR 提取的文本发送给 LLM 进行结构化提取（发送前自动脱敏）。

    Args:
        ocr_text: OCR 识别出的原始文本
        ai_prompt: 提取用的 Prompt

    Returns:
        tuple: (parsed_dict, raw_text)
    """
    masked_text, _report = desensitize_text(ocr_text)
    combined_prompt = (
        ai_prompt
        + "\n\n以下是通过OCR识别出的医疗文档文本，请按上述要求提取结构化信息：\n\n"
        + masked_text
    )
    response = client.chat.completions.create(
        model=MODEL_NAME,
        messages=[{'role': 'user', 'content': combined_prompt}],
        temperature=0.1,
        max_tokens=4096
    )
    raw_text = response.choices[0].message.content
    parsed = parse_ai_response(raw_text)
    return parsed, raw_text

# ========== 结构化数据工具函数 ==========

def extract_nested_field(data, field_path):
    """
    递归提取嵌套字段值，支持点分路径（如 demographics.年龄）。

    Args:
        data: 嵌套字典
        field_path: 点分字段路径字符串

    Returns:
        提取到的值，未找到时返回 None
    """
    parts = field_path.split('.', 1)
    if not isinstance(data, dict):
        return None
    value = data.get(parts[0])
    if len(parts) == 1:
        return value
    return extract_nested_field(value, parts[1])


def collect_field_paths(data, prefix=''):
    """
    递归收集 JSON 中所有叶子字段路径（跳过 confidence 字段）。

    Args:
        data: 嵌套字典
        prefix: 当前路径前缀

    Returns:
        list: 所有叶子字段的点分路径列表
    """
    paths = []
    if isinstance(data, dict):
        for key, value in data.items():
            if key == 'confidence':
                continue
            full_path = f"{prefix}.{key}" if prefix else key
            if isinstance(value, dict):
                paths.extend(collect_field_paths(value, full_path))
            elif isinstance(value, list):
                if value and isinstance(value[0], dict):
                    paths.extend(collect_field_paths(value[0], full_path + '[]'))
                else:
                    paths.append(full_path)
            else:
                paths.append(full_path)
    return paths


def is_numeric(value):
    """判断值是否可转换为数值。"""
    if value is None:
        return False
    try:
        float(value)
        return True
    except (ValueError, TypeError):
        return False

# ========== 统计分析 ==========

def analyze_structured_data(record_ids, fields, analysis_type='descriptive'):
    """
    对选中记录的结构化数据进行统计分析，返回统计量和 ECharts 图表配置。

    Args:
        record_ids: 病历记录 ID 列表
        fields: 要分析的字段路径列表（支持点分路径）
        analysis_type: 分析类型（descriptive / trend / distribution）

    Returns:
        dict: {'statistics': {...}, 'charts': [...]}
    """
    from memory import get_db

    conn = get_db()
    cursor = conn.cursor()
    placeholders = ','.join(['?'] * len(record_ids))
    cursor.execute(
        f'SELECT extracted_data, create_time FROM medical_records WHERE id IN ({placeholders})',
        record_ids
    )
    rows = cursor.fetchall()
    conn.close()

    field_values = {field: [] for field in fields}
    time_series = {field: [] for field in fields}

    for row in rows:
        if not row['extracted_data']:
            continue
        data = json.loads(row['extracted_data'])
        create_time = row['create_time'] or ''
        for field in fields:
            value = extract_nested_field(data, field)
            if is_numeric(value):
                field_values[field].append(float(value))
                time_series[field].append({'time': create_time, 'value': float(value)})

    # 计算统计量
    statistics = {}
    for field, values in field_values.items():
        if not values:
            statistics[field] = {'count': 0, 'msg': '无有效数值'}
            continue
        arr = np.array(values)
        statistics[field] = {
            'count': len(values),
            'mean': round(float(np.mean(arr)), 2),
            'median': round(float(np.median(arr)), 2),
            'std': round(float(np.std(arr)), 2),
            'min': round(float(np.min(arr)), 2),
            'max': round(float(np.max(arr)), 2),
            'q1': round(float(np.percentile(arr, 25)), 2),
            'q3': round(float(np.percentile(arr, 75)), 2)
        }

    chart_configs = []
    valid_fields = [field for field in fields if field_values[field]]

    if analysis_type == 'descriptive' and valid_fields:
        chart_configs.append({
            'title': {'text': '字段均值对比', 'left': 'center'},
            'tooltip': {'trigger': 'axis'},
            'xAxis': {'type': 'category', 'data': [f.split('.')[-1] for f in valid_fields],
                      'axisLabel': {'rotate': 30}},
            'yAxis': {'type': 'value', 'name': '均值'},
            'series': [{
                'data': [statistics[f]['mean'] for f in valid_fields],
                'type': 'bar',
                'itemStyle': {'color': '#2563eb'},
                'label': {'show': True, 'position': 'top'}
            }],
            'grid': {'bottom': 80}
        })

        if len(valid_fields) <= 8:
            boxplot_data = []
            for field in valid_fields:
                stat = statistics[field]
                boxplot_data.append([stat['min'], stat['q1'], stat['median'], stat['q3'], stat['max']])
            chart_configs.append({
                'title': {'text': '数据分布（箱线图）', 'left': 'center'},
                'tooltip': {'trigger': 'item'},
                'xAxis': {'type': 'category', 'data': [f.split('.')[-1] for f in valid_fields],
                          'axisLabel': {'rotate': 30}},
                'yAxis': {'type': 'value'},
                'series': [{
                    'type': 'boxplot',
                    'data': boxplot_data,
                    'itemStyle': {'color': '#dbeafe', 'borderColor': '#2563eb'}
                }],
                'grid': {'bottom': 80}
            })

    elif analysis_type == 'trend' and valid_fields:
        series_list = []
        colors = ['#2563eb', '#059669', '#d97706', '#dc2626', '#7c3aed']
        for index, field in enumerate(valid_fields):
            sorted_ts = sorted(time_series[field], key=lambda item: item['time'])
            series_list.append({
                'name': field.split('.')[-1],
                'type': 'line',
                'data': [item['value'] for item in sorted_ts],
                'smooth': True,
                'itemStyle': {'color': colors[index % len(colors)]}
            })
        all_times = sorted(set(
            item['time'] for field in valid_fields for item in time_series[field]
        ))
        chart_configs.append({
            'title': {'text': '时间趋势分析', 'left': 'center'},
            'tooltip': {'trigger': 'axis'},
            'legend': {'data': [f.split('.')[-1] for f in valid_fields], 'bottom': 0},
            'xAxis': {'type': 'category', 'data': all_times, 'axisLabel': {'rotate': 45}},
            'yAxis': {'type': 'value'},
            'series': series_list,
            'grid': {'bottom': 80}
        })

    elif analysis_type == 'distribution' and valid_fields:
        first_field = valid_fields[0]
        values = field_values[first_field]
        n_bins = min(10, max(3, len(values) // 2))
        hist_counts, bin_edges = np.histogram(values, bins=n_bins)
        bin_labels = [f"{bin_edges[i]:.1f}-{bin_edges[i+1]:.1f}" for i in range(len(hist_counts))]
        chart_configs.append({
            'title': {'text': f'{first_field.split(".")[-1]} 频次分布', 'left': 'center'},
            'tooltip': {'trigger': 'axis'},
            'xAxis': {'type': 'category', 'data': bin_labels, 'axisLabel': {'rotate': 30}},
            'yAxis': {'type': 'value', 'name': '频次'},
            'series': [{
                'data': [int(count) for count in hist_counts],
                'type': 'bar',
                'itemStyle': {'color': '#059669'}
            }],
            'grid': {'bottom': 80}
        })

        if len(valid_fields) == 1:
            pie_data = [
                {'name': bin_labels[i], 'value': int(hist_counts[i])}
                for i in range(len(hist_counts)) if hist_counts[i] > 0
            ]
            chart_configs.append({
                'title': {'text': f'{first_field.split(".")[-1]} 分段占比', 'left': 'center'},
                'tooltip': {'trigger': 'item', 'formatter': '{b}: {c} ({d}%)'},
                'series': [{
                    'type': 'pie',
                    'radius': ['40%', '70%'],
                    'data': pie_data,
                    'label': {'formatter': '{b}\n{d}%'}
                }]
            })

    return {'statistics': statistics, 'charts': chart_configs}

# ========== Excel 导出 ==========

def generate_excel(data_list, upload_folder):
    """
    多角色 Excel 导出，不同角色放不同 Sheet，低置信度字段标红。

    Args:
        data_list: 病历记录列表（每项含 role_id、extracted_data、confidence_data 等）
        upload_folder: 输出文件目录

    Returns:
        str: 生成的 Excel 文件路径
    """
    from openpyxl.styles import Font, PatternFill
    red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    red_font = Font(color='CC0000', bold=True)

    # 按角色分组
    grouped = {}
    for item in data_list:
        role = item.get('role_id', 'other')
        role_name = CATEGORY_CONFIGS.get(role, {}).get('name', role)
        grouped.setdefault(role_name, []).append(item)

    excel_path = os.path.join(
        upload_folder,
        f"临床数据_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        for sheet_name, items in grouped.items():
            rows = []
            low_confidence_fields = set()

            for item in items:
                row = {
                    '病历编号': item.get('case_number', ''),
                    '模板': item.get('template_name', ''),
                    '录入时间': item.get('create_time', ''),
                    '数据来源': '录音' if item.get('source_type') == 'audio' else '图片',
                }
                data = item.get('extracted_data', {})
                confidence = item.get('confidence_data', {})

                _flatten_to_row(data, row, '', confidence, low_confidence_fields)

                if item.get('source_type') == 'audio' and item.get('audio_transcript'):
                    transcript = item['audio_transcript']
                    row['转录原文'] = transcript[:500] + ('...' if len(transcript) > 500 else '')

                qualitative_data = item.get('qualitative_data')
                if qualitative_data and isinstance(qualitative_data, dict):
                    row['主题分析'] = ', '.join(qualitative_data.get('themes', []))
                    row['关键词'] = ', '.join(qualitative_data.get('keywords', []))
                    row['情感倾向'] = qualitative_data.get('sentiment', '')

                rows.append(row)

            if not rows:
                continue

            df = pd.DataFrame(rows)
            safe_sheet_name = sheet_name[:31]
            df.to_excel(writer, index=False, sheet_name=safe_sheet_name)

            if low_confidence_fields:
                worksheet = writer.sheets[safe_sheet_name]
                for col_idx, col_name in enumerate(df.columns, 1):
                    for field_name in low_confidence_fields:
                        if field_name in col_name:
                            for row_idx in range(2, len(df) + 2):
                                cell = worksheet.cell(row=row_idx, column=col_idx)
                                cell.fill = red_fill
                                cell.font = red_font
                            worksheet.cell(row=1, column=col_idx).fill = red_fill
                            worksheet.cell(row=1, column=col_idx).font = red_font
                            break

    return excel_path


def _flatten_to_row(data, row, prefix, confidence, low_confidence_fields):
    """递归展平嵌套 JSON 为 Excel 行的列。"""
    if isinstance(data, dict):
        for key, value in data.items():
            if key == 'confidence':
                _collect_low_confidence(value, low_confidence_fields)
                continue
            full_key = f"{prefix}_{key}" if prefix else key
            if isinstance(value, dict):
                _flatten_to_row(value, row, full_key, confidence, low_confidence_fields)
            elif isinstance(value, list):
                _flatten_list_to_row(value, row, full_key, confidence, low_confidence_fields)
            else:
                row[full_key] = value
    elif isinstance(data, list):
        _flatten_list_to_row(data, row, prefix, confidence, low_confidence_fields)


def _flatten_list_to_row(data_list, row, prefix, confidence, low_confidence_fields):
    """展平列表数据为 Excel 行的列。"""
    for index, item in enumerate(data_list):
        if isinstance(item, dict):
            name_key = (
                item.get('项目名称')
                or item.get('英文缩写')
                or item.get('项目')
                or item.get('诊断名称')
                or str(index + 1)
            )
            item_prefix = f"{prefix}_{name_key}" if prefix else name_key
            for key, value in item.items():
                if key in ('项目名称', '英文缩写', '项目', '诊断名称'):
                    continue
                if isinstance(value, (dict, list)):
                    continue
                col = f"{item_prefix}_{key}" if key not in ('数值', '评分') else item_prefix
                row[col] = value
        elif isinstance(item, str):
            row[f"{prefix}_{index + 1}"] = item


def _collect_low_confidence(confidence_data, low_confidence_fields):
    """收集置信度低于 0.9 的字段名。"""
    if isinstance(confidence_data, dict):
        for key, value in confidence_data.items():
            if isinstance(value, dict):
                _collect_low_confidence(value, low_confidence_fields)
            else:
                try:
                    if float(value) < 0.9:
                        low_confidence_fields.add(key)
                except (ValueError, TypeError):
                    pass
